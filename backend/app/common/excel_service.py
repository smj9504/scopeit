"""
ScopeIt - Excel Import/Export Service
Shared service for generating Excel templates and parsing uploaded Excel files
for both estimates and invoices.
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
from typing import Literal
from decimal import Decimal, InvalidOperation
from fastapi import HTTPException


# Column definitions: (header_label, key, column_width)
COLUMNS = [
    ("Section", "section", 20),
    ("Code", "code", 15),
    ("Name", "name", 30),
    ("Description", "description", 40),
    ("Unit", "unit", 10),
    ("Quantity", "quantity", 12),
    ("Unit Price", "unit_price", 14),
    ("Taxable (Yes/No)", "taxable", 16),
]

VALID_UNITS = ["EA", "SF", "LF", "SQ", "HR", "DAY", "LOT"]


def generate_template(document_type: Literal["estimate", "invoice"]) -> BytesIO:
    """Generate an Excel template for estimate or invoice import."""
    wb = Workbook()

    # Instructions sheet
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"
    _build_instructions_sheet(ws_instructions, document_type)

    # Items sheet
    ws_items = wb.create_sheet("Items")
    _build_items_sheet(ws_items)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def parse_excel_file(file_bytes: bytes) -> dict:
    """
    Parse an uploaded Excel file and return structured data.

    Returns:
        {
            "sections": [{
                "name": str,
                "order_index": int,
                "items": [{
                    "code": str|None, "name": str, "description": str|None,
                    "unit": str|None, "quantity": float, "unit_price": float,
                    "is_taxable": bool, "order_index": int
                }]
            }],
            "errors": [str],
            "warnings": [str],
            "total_items": int,
        }
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Invalid Excel file. Please upload a valid .xlsx file.",
        )

    items_sheet = _find_items_sheet(wb)
    return _parse_items_sheet(items_sheet)


def _find_items_sheet(wb):
    """Find the Items sheet in the workbook."""
    for name in wb.sheetnames:
        if name.lower() == "items":
            return wb[name]

    # Fallback: use the only data sheet
    non_instruction = [s for s in wb.sheetnames if s.lower() != "instructions"]
    if len(non_instruction) == 1:
        return wb[non_instruction[0]]
    if len(wb.sheetnames) == 1:
        return wb.active

    raise HTTPException(
        status_code=400,
        detail="Could not find 'Items' sheet in the uploaded file.",
    )


def _build_instructions_sheet(ws, document_type: str):
    """Populate the Instructions sheet."""
    header_font = Font(bold=True, size=14, color="111827")
    subheader_font = Font(bold=True, size=11, color="374151")
    body_font = Font(size=11, color="4B5563")

    ws.column_dimensions["A"].width = 80

    instructions = [
        (f"ScopeIt - {document_type.title()} Import Template", header_font),
        ("", body_font),
        ("How to use this template:", subheader_font),
        ("1. Go to the 'Items' sheet (tab at the bottom)", body_font),
        ("2. Fill in your line items row by row", body_font),
        ("3. Use the 'Section' column to group items into sections", body_font),
        ("   - Items with the same Section name are grouped together", body_font),
        ("   - If left blank, items go into a 'General' section", body_font),
        ("4. Save the file and upload it in ScopeIt", body_font),
        ("", body_font),
        ("Column Reference:", subheader_font),
        ("Section    - Group name for organizing items (e.g., 'Demolition', 'Framing')", body_font),
        ("Code       - Optional item code (e.g., 'DEM-001')", body_font),
        ("Name       - Item name (REQUIRED)", body_font),
        ("Description - Detailed description of the work", body_font),
        ("Unit       - Unit of measure: EA, SF, LF, SQ, HR, DAY, LOT", body_font),
        ("Quantity   - Number of units (default: 1)", body_font),
        ("Unit Price - Price per unit in dollars (default: 0)", body_font),
        ("Taxable    - Yes or No (default: Yes)", body_font),
        ("", body_font),
        ("Tips:", subheader_font),
        ("- The Name column is the only required field", body_font),
        ("- Delete the example rows before importing", body_font),
        ("- You can add as many rows as needed", body_font),
    ]

    for i, (text, font) in enumerate(instructions, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = font


def _build_items_sheet(ws):
    """Build the Items data sheet with headers, styling, and validation."""
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(bottom=Side(style="thin", color="E5E7EB"))

    # Write headers
    for col_idx, (label, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validation: Taxable (Yes/No)
    taxable_col = len(COLUMNS)
    dv_taxable = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_taxable.error = "Please enter Yes or No"
    dv_taxable.errorTitle = "Invalid Value"
    ws.add_data_validation(dv_taxable)
    dv_taxable.add(f"{get_column_letter(taxable_col)}2:{get_column_letter(taxable_col)}1000")

    # Data validation: Unit
    unit_col = 5
    units_str = ",".join(VALID_UNITS)
    dv_unit = DataValidation(type="list", formula1=f'"{units_str}"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add(f"{get_column_letter(unit_col)}2:{get_column_letter(unit_col)}1000")

    # Example rows
    examples = [
        ("Demolition", "DEM-001", "Remove existing flooring", "Remove and dispose of damaged carpet and pad", "SF", 500, 2.50, "Yes"),
        ("Demolition", "DEM-002", "Remove baseboards", "Carefully remove baseboards for reinstallation", "LF", 120, 1.75, "Yes"),
        ("Framing", "FRM-001", "Replace wall studs", "Replace damaged 2x4 wall studs", "EA", 8, 45.00, "Yes"),
        ("Paint", "PNT-001", "Prime and paint walls", "2 coats primer, 2 coats paint", "SF", 800, 3.25, "No"),
    ]

    body_font = Font(size=11, color="374151")
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.border = thin_border

    ws.freeze_panes = "A2"


def _parse_items_sheet(ws) -> dict:
    """Parse the items sheet into structured section/item data."""
    errors = []
    warnings = []
    sections_dict = {}
    total_items = 0

    # Find header row by looking for "Name" column
    header_row = None
    col_map = {}
    for row_idx in range(1, 6):
        for col_idx in range(1, 20):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip().lower() == "name":
                header_row = row_idx
                break
        if header_row:
            break

    if header_row is None:
        raise HTTPException(
            status_code=400,
            detail="Could not find header row. Make sure the 'Name' column header exists.",
        )

    # Map column indices from header
    for col_idx in range(1, 20):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            normalized = (
                str(val).strip().lower()
                .replace(" ", "_")
                .replace("(yes/no)", "")
                .strip("_")
            )
            col_map[normalized] = col_idx

    if "name" not in col_map:
        raise HTTPException(status_code=400, detail="Missing required 'Name' column.")

    max_col = max(col_map.values())

    # Parse data rows
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        name_val = ws.cell(row=row_idx, column=col_map["name"]).value
        if not name_val or str(name_val).strip() == "":
            errors.append(f"Row {row_idx}: Missing required 'Name' field, skipping.")
            continue

        section_name = _get_cell_str(ws, row_idx, col_map.get("section")) or "General"
        code = _get_cell_str(ws, row_idx, col_map.get("code"))
        name = str(name_val).strip()
        description = _get_cell_str(ws, row_idx, col_map.get("description"))
        unit = _get_cell_str(ws, row_idx, col_map.get("unit"))

        quantity = _get_cell_decimal(ws, row_idx, col_map.get("quantity"), default=Decimal("1"))
        if quantity is None:
            errors.append(f"Row {row_idx}: Invalid quantity, using 1.")
            quantity = Decimal("1")

        unit_price = _get_cell_decimal(ws, row_idx, col_map.get("unit_price"), default=Decimal("0"))
        if unit_price is None:
            errors.append(f"Row {row_idx}: Invalid unit price, using 0.")
            unit_price = Decimal("0")

        is_taxable = _get_cell_bool(ws, row_idx, col_map.get("taxable"), default=True)

        item = {
            "code": code,
            "name": name,
            "description": description,
            "unit": unit,
            "quantity": float(quantity),
            "unit_price": float(unit_price),
            "is_taxable": is_taxable,
        }

        if section_name not in sections_dict:
            sections_dict[section_name] = []
        sections_dict[section_name].append(item)
        total_items += 1

    if total_items == 0:
        raise HTTPException(
            status_code=400,
            detail="No valid items found in the Excel file. Please check the format and try again.",
        )

    # Build sections list with order_index
    sections = []
    for idx, (section_name, items) in enumerate(sections_dict.items()):
        for item_idx, item in enumerate(items):
            item["order_index"] = item_idx
        sections.append({
            "name": section_name,
            "order_index": idx,
            "items": items,
        })

    return {
        "sections": sections,
        "errors": errors,
        "warnings": warnings,
        "total_items": total_items,
    }


# --- Helpers ---

def _get_cell_str(ws, row, col):
    if col is None:
        return None
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    return str(val).strip() or None


def _get_cell_decimal(ws, row, col, default=None):
    if col is None:
        return default
    val = ws.cell(row=row, column=col).value
    if val is None:
        return default
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return None


def _get_cell_bool(ws, row, col, default=True):
    if col is None:
        return default
    val = ws.cell(row=row, column=col).value
    if val is None:
        return default
    s = str(val).strip().lower()
    if s in ("no", "false", "0", "n"):
        return False
    return True
