"""
ScopeIt - Packing Tool Export Service

Ported from moving_estimate/moving-estimator-backend/services/export.py.
Generates professional PDF and Excel exports for packing/contents estimates.

Original source: moving_estimate moving-estimator-backend export service
(SOS Moving Company estimate format).
"""

import io
import re
from datetime import datetime
from typing import Dict, Any, Optional, List
import random
import string

# PDF Generation
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, HRFlowable, KeepTogether, BaseDocTemplate, Frame, PageTemplate
)
from reportlab.pdfgen import canvas

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================
# SCOPEIT COMPANY INFO HELPER
# ============================================

def build_company_info(company, override=None) -> dict:
    """Build company info dict from ScopeIt Company model with optional overrides.

    ``override`` may be a dict or a Pydantic model instance.
    """
    info = {
        "name": company.name if company else "",
        "address": f"{company.address_line1 or ''}, {company.city or ''}, {company.state or ''} {company.zipcode or ''}".strip(", ") if company else "",
        "phone": company.phone if company else "",
        "email": company.email if company else "",
    }
    if override:
        # Handle both dict and Pydantic model
        if hasattr(override, "model_dump"):
            override = override.model_dump()
        elif not isinstance(override, dict):
            override = dict(override)
        for key in ("name", "address", "phone", "email", "license"):
            val = override.get(key)
            if val:
                info[key] = val
    return info


# ============================================
# DEFAULT PRICES
# ============================================
DEFAULT_PRICES = {
    'labor': 68.00, 'labor_fragile': 83.00, 'labor_specialty': 125.00, 'supervisor': 83.00,
    'box_small': 2.95, 'box_medium': 3.91, 'box_large': 5.28, 'box_xlarge': 6.40,
    'box_dish': 9.98, 'box_wardrobe': 18.48, 'box_mirror': 10.29, 'box_tv': 30.24,
    'blanket': 18.26, 'bubble_12': 23.98, 'bubble_antistatic': 42.00,
    'packing_paper': 18.00, 'shrink_wrap': 29.83,
    'poly_bags': 24.00, 'furniture_bags': 8.57, 'appliance_tape': 29.90,
    'packing_tape': 8.59, 'inventory_tags': 60.11, 'box_liners': 6.50,
    'truck_26': 197.00, 'storage_sf': 2.20,
    'waste_removal': 67.63,
}


def generate_estimate_number():
    """Generate a unique estimate number.

    Note: Used as a fallback only. In ScopeIt, estimate numbers are normally
    generated via ``company.estimate_prefix``.
    """
    chars = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    year = datetime.now().year
    seq = random.randint(1, 9999)
    return f"EST-{chars}-{year}-{seq:04d}"


class DescriptionBuilder:
    """Builds dynamic line item descriptions from room summaries."""

    def __init__(self, room_summaries=None):
        self.summaries = room_summaries or []
        self._notable = []
        self._high_value = []
        self._categories = set()
        self._packing_notes = []
        self._room_names = []
        for s in self.summaries:
            rs = s if isinstance(s, dict) else s.dict()
            self._notable.extend(rs.get("notable_items", []))
            self._high_value.extend(rs.get("high_value_items", []))
            self._categories.update(rs.get("categories_present", []))
            self._packing_notes.extend(rs.get("packing_notes", []))
            self._room_names.append(rs.get("room_name", ""))

    @property
    def has_context(self):
        return len(self.summaries) > 0

    @staticmethod
    def _list_str(items, max_items=3):
        """Format list: 'A, B, and C'."""
        items = [i for i in items[:max_items] if i]
        if not items:
            return ""
        if len(items) == 1:
            return items[0]
        return ", ".join(items[:-1]) + ", and " + items[-1]

    def _rooms_str(self, max_rooms=3):
        return self._list_str(self._room_names, max_rooms)

    # ---- Labor descriptions ----

    def pack_out_crew(self, crew_size, fallback):
        if not self.has_context:
            return fallback
        rooms = self._rooms_str()
        notable = self._list_str(self._notable, 4)
        if notable:
            return (
                f'{crew_size}-person crew, professional packing '
                f'of {rooms} contents including {notable}'
            )
        cats = self._list_str(sorted(self._categories), 4)
        if cats:
            return (
                f'{crew_size}-person crew, packing {rooms} '
                f'— {cats}'
            )
        return fallback

    def supervisor(self, fallback):
        if not self.has_context or not self._room_names:
            return fallback
        return (
            f'On-site supervision across {len(self._room_names)} '
            f'rooms, inventory documentation, quality control'
        )

    def specialized_handling(self, fallback):
        if self._high_value:
            items = self._list_str(self._high_value, 3)
            return f'Custom handling for high-value items: {items}'
        specialty = [n for n in self._notable
                     if any(c in self._categories
                            for c in ("Electronics", "Fragile",
                                      "Artwork", "Collectibles"))]
        if specialty:
            items = self._list_str(specialty, 3)
            return (
                f'Extra care packaging: {items} — '
                f'includes custom crating as needed'
            )
        return fallback

    def pack_back_crew(self, crew_size, fallback):
        if not self.has_context:
            return fallback
        rooms = self._rooms_str()
        return (
            f'{crew_size}-person crew, unloading and placement '
            f'in {rooms}, unpacking'
        )

    def pack_back_specialized(self, fallback):
        if self._high_value:
            items = self._list_str(self._high_value, 3)
            return (
                f'Careful unpacking and placement: {items}'
            )
        return fallback

    # ---- Material descriptions ----

    def _items_in_category(self, *cats):
        """Find notable items from rooms that have given categories."""
        result = []
        for s in self.summaries:
            rs = s if isinstance(s, dict) else s.dict()
            room_cats = set(rs.get("categories_present", []))
            if room_cats & set(cats):
                result.extend(rs.get("notable_items", []))
        return result

    def material_desc(self, mat_key, fallback):
        """Return generic material description without listing content items."""
        return fallback


def get_line_items_from_estimate(
    estimate_data: Dict[str, Any],
    prices: Optional[Dict[str, float]] = None,
) -> List[Dict]:
    """Convert estimate data to detailed line items matching SOS format.

    Args:
        estimate_data: The estimate dict (from calculator response).
        prices: Optional price dict from DB. Falls back to DEFAULT_PRICES.
    """
    P = {**DEFAULT_PRICES, **(prices or {})}

    total_rooms = estimate_data.get('total_rooms', estimate_data.get('rooms', 0))
    total_items = estimate_data.get('total_items', estimate_data.get('items', 0))
    total_hours = estimate_data.get('total_hours', estimate_data.get('hours', 0))
    crew_size = estimate_data.get('crew_size', estimate_data.get('crew', 4))
    materials = estimate_data.get('materials', {})
    include_packback = estimate_data.get('include_packback', True)
    storage_months = estimate_data.get('storage_months', 1)
    storage_sf = estimate_data.get('storage_sf', 0)
    staging_type = estimate_data.get('staging_type', 'off_site')
    is_on_site = staging_type == 'on_site'
    # Dynamic description builder from room summaries
    room_summaries = estimate_data.get('room_summaries', [])
    desc = DescriptionBuilder(room_summaries)

    pack_out_hours = total_hours * 0.55 if include_packback else total_hours * 0.8
    pack_back_hours = total_hours * 0.45 if include_packback else 0

    sections = []

    # ========== SECTION 1: PACK-OUT LABOR SERVICES ==========
    packout_items = []

    # Calculate hour allocations (kept for internal ratio)
    std_hours = max(2, round(pack_out_hours * 0.6))
    fragile_hours = max(1, round(pack_out_hours * 0.15))
    specialty_hours = max(1, round(pack_out_hours * 0.08))
    furniture_hours = max(1, round(pack_out_hours * 0.1))
    appliance_hours = max(1, round(pack_out_hours * 0.08))
    inventory_hours = max(1, round(total_hours * 0.05))
    supervisor_hours = max(1, round(total_hours * 0.10))

    # 1. Pack-Out Crew Labor (standard + furniture + appliance + inventory)
    crew_hours = std_hours + furniture_hours + appliance_hours + inventory_hours
    packout_items.append({
        'name': 'Pack-Out Crew Labor',
        'detail': desc.pack_out_crew(
            crew_size,
            f'{crew_size}-person crew, professional packing of all contents '
            'including wrapping, boxing, labeling, and loading'
        ),
        'qty': crew_hours, 'unit': 'HR', 'price': P['labor']
    })

    # 2. Supervisor/Foreman
    packout_items.append({
        'name': 'Supervisor/Foreman',
        'detail': desc.supervisor(
            'On-site supervision, inventory documentation, '
            'quality control, and client communication'
        ),
        'qty': supervisor_hours, 'unit': 'HR', 'price': P['supervisor']
    })

    # 3. Specialized Handling (fragile + specialty combined)
    specialized_hours = fragile_hours + specialty_hours
    packout_items.append({
        'name': 'Specialized Handling',
        'detail': desc.specialized_handling(
            'Electronics, fragile items, artwork - includes extra care '
            'packaging and custom crating as needed'
        ),
        'qty': specialized_hours, 'unit': 'HR', 'price': P['labor_specialty']
    })

    sections.append({'title': 'PACK-OUT LABOR SERVICES', 'items': packout_items})

    # ========== SECTION 2: PACKING MATERIALS ==========
    material_items = []

    # Boxes
    if materials.get('box_small', 0) > 0:
        material_items.append({
            'name': 'Small Cartons (1.5 Cu Ft)',
            'detail': desc.material_desc(
                'box_small',
                'Books, small decor, electronics — includes packing paper & tape'
            ),
            'qty': materials['box_small'], 'unit': 'EA', 'price': P['box_small']
        })

    if materials.get('box_medium', 0) > 0:
        material_items.append({
            'name': 'Medium Cartons (3.0 Cu Ft)',
            'detail': desc.material_desc(
                'box_medium',
                'General household, consumables — includes packing paper & tape'
            ),
            'qty': materials['box_medium'], 'unit': 'EA', 'price': P['box_medium']
        })

    if materials.get('box_large', 0) > 0:
        material_items.append({
            'name': 'Large Cartons (4.5 Cu Ft)',
            'detail': '',
            'qty': materials['box_large'], 'unit': 'EA', 'price': P['box_large']
        })

    if materials.get('box_dish', 0) > 0:
        material_items.append({
            'name': 'Dish-Pack / Reinforced Cartons',
            'detail': desc.material_desc(
                'box_dish',
                'Frames, fragile decor, glass — includes packing paper & tape'
            ),
            'qty': materials['box_dish'], 'unit': 'EA', 'price': P['box_dish']
        })

    if materials.get('box_mirror', 0) > 0:
        material_items.append({
            'name': 'Telescopic Picture/Mirror Boxes',
            'detail': desc.material_desc(
                'box_mirror',
                'Large framed art & mirrors — includes packing paper & tape'
            ),
            'qty': materials['box_mirror'], 'unit': 'EA', 'price': P['box_mirror']
        })

    if materials.get('box_wardrobe', 0) > 0:
        material_items.append({
            'name': 'Wardrobe Boxes',
            'detail': desc.material_desc('box_wardrobe', ''),
            'qty': materials['box_wardrobe'], 'unit': 'EA',
            'price': P['box_wardrobe']
        })

    # Protective materials
    if materials.get('blanket', 0) > 0:
        material_items.append({
            'name': 'Heavy-Duty Moving Pads (72" × 80")',
            'detail': desc.material_desc(
                'blanket',
                'Furniture wrapping — bed frames, tables, dressers, appliances'
            ),
            'qty': materials['blanket'], 'unit': 'EA', 'price': P['blanket']
        })

    if materials.get('shrink_wrap', 0) > 0:
        material_items.append({
            'name': '4-Mil Stretch Wrap — 20" × 1000\' Rolls',
            'detail': '',
            'qty': materials['shrink_wrap'], 'unit': 'RL', 'price': P['shrink_wrap']
        })

    if materials.get('bubble_12', 0) > 0:
        material_items.append({
            'name': '1/2" Professional Bubble Wrap — 12" × 250\' Rolls',
            'detail': desc.material_desc(
                'bubble_12',
                'Vanity mirrors, electronics, art pieces — economy grade'
            ),
            'qty': materials['bubble_12'], 'unit': 'RL', 'price': P['bubble_12']
        })

    # Anti-static bubble
    material_items.append({
        'name': 'Anti-Static Bubble Wrap — Electronics Protection',
        'detail': 'Baby monitor, sound machine, lamps',
        'qty': 1, 'unit': 'RL', 'price': P['bubble_antistatic']
    })

    if materials.get('packing_paper', 0) > 0:
        material_items.append({
            'name': 'Packing Paper — Bundle (50 lb)',
            'detail': 'Wrapping dishes, glassware, and fragile items',
            'qty': materials['packing_paper'], 'unit': 'BN', 'price': P['packing_paper']
        })

    # Poly bags
    material_items.append({
        'name': '6-Mil Clear Poly Bags — Various Sizes',
        'detail': 'Soft goods, bedding, textiles — labeled for restoration - pack/25',
        'qty': 2, 'unit': 'PK', 'price': P['poly_bags']
    })

    # Furniture bags
    material_items.append({
        'name': 'Sealed Sanitary Plastic Furniture Bags',
        'detail': '',
        'qty': 3, 'unit': 'EA', 'price': P['furniture_bags']
    })

    # Appliance tape
    material_items.append({
        'name': 'Non-Residue Tape — Appliance Door Securing',
        'detail': '',
        'qty': 4, 'unit': 'RL', 'price': P['appliance_tape']
    })

    # Packing tape
    material_items.append({
        'name': 'Packing Tape — Standard (2" × 60 yd)',
        'detail': '',
        'qty': max(1, total_rooms // 5), 'unit': 'RL', 'price': P['packing_tape']
    })

    # Labeling supplies
    material_items.append({
        'name': 'Labeling Supplies — Pre-printed labels, markers, inventory tags',
        'detail': '',
        'qty': max(2, total_rooms // 4), 'unit': 'KT', 'price': P['inventory_tags']
    })

    # Box liners
    material_items.append({
        'name': 'Plastic-Lined Leak-Prevention Box Liners (Medium)',
        'detail': '',
        'qty': max(4, total_rooms // 2), 'unit': 'EA', 'price': P['box_liners']
    })

    sections.append({'title': 'PACKING MATERIALS & SUPPLIES', 'items': material_items})

    # ========== SECTION 3: TRANSPORTATION / ON-SITE RELOCATION ==========
    if is_on_site:
        on_site_fee = crew_size * 1.0 * P['labor']
        sections.append({
            'title': 'ON-SITE CONTENT RELOCATION',
            'items': [{
                'name': 'On-Site Content Relocation — Pack-Out',
                'detail': f'{crew_size}-person crew moving contents to designated staging area within property',
                'qty': 1, 'unit': 'EA', 'price': round(on_site_fee, 2)
            }]
        })
    else:
        sections.append({
            'title': 'TRANSPORTATION & LOGISTICS',
            'items': [{
                'name': 'Moving Van (21\'–27\') and Equipment — Pack-Out Trip',
                'detail': 'Residence to storage facility, loaded — per day rate',
                'qty': 1, 'unit': 'DY', 'price': P['truck_26']
            }]
        })

        # ========== SECTION 4: STORAGE ==========
        if storage_months > 0:
            storage_items = []
            sf_rate = P.get('storage_sf', 2.20)
            # Guard: storage sf rate must be a realistic per-SF/mo value (not a total cost)
            if not (1.0 <= sf_rate <= 10.0):
                sf_rate = 2.20
            # Snap to standard storage unit sizes (5x5, 5x10, 10x10, etc.)
            STANDARD_UNITS = [25, 50, 100, 150, 200, 250, 300]
            storage_sf = max(25, storage_sf)
            for unit in STANDARD_UNITS:
                if storage_sf <= unit:
                    storage_sf = unit
                    break
            else:
                storage_sf = STANDARD_UNITS[-1]
            monthly_cost = storage_sf * sf_rate
            storage_items.append({
                'name': 'Climate-Controlled Off-Site Storage & Insurance',
                'detail': (
                    f'{storage_sf} SF required; temp 55\u201380\u00b0F; '
                    f'humidity controlled; ${sf_rate:.2f}/SF/mo'
                ),
                'qty': storage_months, 'unit': 'MO',
                'price': round(monthly_cost, 2)
            })
            # Setup fee scales with unit size (power-law ^0.65, base $85 for 10x10)
            SETUP_BY_SIZE = {25: 42, 50: 54, 100: 85, 150: 109, 200: 131, 250: 152, 300: 172}
            setup_fee = SETUP_BY_SIZE.get(storage_sf, 85)
            storage_items.append({
                'name': 'Initial Storage Setup — Unit Preparation & Organization',
                'detail': f'{storage_sf} SF unit — shelving, inventory placement, padlock',
                'qty': 1, 'unit': 'EA', 'price': setup_fee
            })
            sections.append({'title': 'CLIMATE-CONTROLLED STORAGE', 'items': storage_items})

    # ========== SECTION 5: DEBRIS HAULING ==========
    total_material_qty = sum(materials.values()) if materials else 0
    if total_material_qty > 0:
        debris_hours = max(1, total_material_qty // 40 + 1)
        sections.append({
            'title': 'DEBRIS HAULING & DISPOSAL',
            'items': [{
                'name': 'Debris Hauling — Packing Waste & Damaged Materials',
                'detail': 'Removal and disposal of damaged contents, '
                          'packing debris, and non-salvageable items',
                'qty': debris_hours, 'unit': 'HR',
                'price': P['waste_removal']
            }]
        })

    # ========== SECTION 6: PACK-BACK LABOR SERVICES ==========
    if include_packback:
        packback_items = []

        # Calculate pack-back hour allocations
        pb_crew_base = max(2, round(pack_back_hours * 0.65))
        pb_reassembly = max(1, round(pack_back_hours * 0.12))
        pb_appliance = max(1, round(pack_back_hours * 0.06))
        pb_waste = max(1, round(pack_back_hours * 0.06))
        pb_supervisor = max(1, round(pack_back_hours * 0.10))

        # 1. Pack-Back Crew Labor (crew + reassembly + appliance + waste)
        pb_total_crew = pb_crew_base + pb_reassembly + pb_appliance + pb_waste
        packback_items.append({
            'name': 'Pack-Back Crew Labor',
            'detail': desc.pack_back_crew(
                crew_size,
                f'{crew_size}-person crew, unloading, placement, '
                'furniture reassembly, and unpacking'
            ),
            'qty': pb_total_crew, 'unit': 'HR', 'price': P['labor']
        })

        # 2. Supervisor/Foreman — Pack-Back
        packback_items.append({
            'name': 'Supervisor/Foreman',
            'detail': 'Pack-back oversight, quality control, client walkthrough',
            'qty': pb_supervisor, 'unit': 'HR', 'price': P['supervisor']
        })

        # 3. Specialized Handling — Pack-Back (same ratio as pack-out)
        spec_ratio = specialized_hours / max(1, crew_hours + specialized_hours)
        pb_specialized = max(1, round(pb_total_crew * spec_ratio))
        packback_items.append({
            'name': 'Specialized Handling',
            'detail': desc.pack_back_specialized(
                'Electronics, fragile items, artwork - '
                'careful unpacking and placement'
            ),
            'qty': pb_specialized, 'unit': 'HR', 'price': P['labor_specialty']
        })

        # Return trip or on-site move-back
        if is_on_site:
            on_site_fee = crew_size * 1.0 * P['labor']
            packback_items.append({
                'name': 'On-Site Content Relocation — Pack-Back',
                'detail': f'{crew_size}-person crew moving contents from staging area back to restored rooms',
                'qty': 1, 'unit': 'EA', 'price': round(on_site_fee, 2)
            })
        else:
            packback_items.append({
                'name': 'Moving Van — Return Trip',
                'detail': 'Storage facility to restored residence — per day rate',
                'qty': 1, 'unit': 'DY', 'price': P['truck_26']
            })

        sections.append({'title': 'PACK-BACK LABOR SERVICES', 'items': packback_items})

    return sections


# ============================================
# SECTION DETAILS → LINE ITEMS CONVERTER
# ============================================

def _section_details_to_line_items(
    section_details: Dict[str, Any],
    sections_totals: Dict[str, float],
    material_details: Optional[List[Dict]] = None,
) -> List[Dict]:
    """Convert section_details (from calculate_estimate_from_content) to the
    line_items format expected by the PDF/Excel export.

    This ensures the export matches the Estimate Editor UI exactly, using the
    same line items and amounts rather than recalculating from totals.
    """
    result = []

    # Section order matching the UI
    section_order = [
        'Pack-Out Labor', 'Materials', 'On-Site Relocation',
        'Transport Out', 'Storage', 'Debris Hauling',
        'Pack-Back Labor', 'Furniture Assembly',
        'Transport Back', 'On-Site Pack-Back Move', 'Special Items',
    ]

    processed = set()
    for section_name in section_order:
        # Materials section: use material_details (not in section_details)
        if section_name == 'Materials' and material_details:
            processed.add(section_name)
            items = []
            for m in material_details:
                items.append({
                    'name': m.get('name', ''),
                    'detail': m.get('detail', ''),
                    'qty': m.get('quantity', 1),
                    'unit': m.get('unit', 'EA'),
                    'price': m.get('unit_price', 0),
                })
            if items:
                result.append({'title': section_name, 'items': items})
            continue

        # Standard sections: use lines from section_details
        detail = section_details.get(section_name)
        if not detail:
            continue
        processed.add(section_name)
        lines = detail.get('lines', [])
        if not lines:
            continue

        items = []
        for line in lines:
            items.append({
                'name': line.get('name', ''),
                'detail': line.get('detail', ''),
                'qty': line.get('qty', 1),
                'unit': line.get('unit', 'HR'),
                'price': line.get('rate', 0),
            })
        result.append({'title': section_name, 'items': items})

    # Any remaining sections not in the predefined order
    for section_name, detail in section_details.items():
        if section_name in processed:
            continue
        lines = detail.get('lines', [])
        if not lines:
            continue
        items = []
        for line in lines:
            items.append({
                'name': line.get('name', ''),
                'detail': line.get('detail', ''),
                'qty': line.get('qty', 1),
                'unit': line.get('unit', 'HR'),
                'price': line.get('rate', 0),
            })
        result.append({'title': section_name, 'items': items})

    return result


# ============================================
# PDF EXPORT - EXACT SOS FORMAT
# ============================================

def generate_estimate_pdf(
    estimate_data: Dict[str, Any],
    client_name: Optional[str] = None,
    client_phone: Optional[str] = None,
    client_email: Optional[str] = None,
    property_address: Optional[str] = None,
    notes: Optional[str] = None,
    company_info: Optional[Dict] = None,
    estimate_number: Optional[str] = None,
    tax_rate: float = 0,
    area_breakdown: Optional[str] = None,
    prices: Optional[Dict[str, float]] = None,
) -> bytes:
    """Generate professional PDF estimate matching SOS Moving format exactly"""

    if not estimate_number:
        estimate_number = generate_estimate_number()

    # Fall back to address embedded in estimate_data (from Photo AI tab)
    if not property_address:
        property_address = estimate_data.get('property_address') or ''

    buffer = io.BytesIO()

    # Priority: 1) frontend line_items (editor edits), 2) section_details from
    # calculate_estimate_from_content, 3) recalculate from totals (legacy fallback).
    # Using section_details ensures PDF/Excel matches the Estimate Editor exactly.
    line_items = estimate_data.get('line_items') or []
    section_details = estimate_data.get('section_details') or {}
    if line_items:
        sections = line_items
    elif section_details:
        sections = _section_details_to_line_items(
            section_details,
            estimate_data.get('sections', {}),
            estimate_data.get('material_details'),
        )
    else:
        sections = get_line_items_from_estimate(estimate_data, prices=prices)

    include_op = estimate_data.get('include_op', False)
    op_rate = estimate_data.get('op_rate', 10)
    include_contingency = estimate_data.get('include_contingency', False)
    contingency_rate = estimate_data.get('contingency_rate', 5)
    subtotal = sum(
        item['qty'] * item['price']
        for section in sections
        for item in section['items']
    )
    # Respect the include flags — if disabled, force 0 regardless of saved value
    op_amount = 0
    if include_op:
        saved_op = estimate_data.get('op_amount')
        op_amount = saved_op if saved_op is not None else subtotal * (op_rate / 100)
    contingency_amount = 0
    if include_contingency:
        saved_cont = estimate_data.get('contingency_amount')
        contingency_amount = saved_cont if saved_cont is not None else subtotal * (contingency_rate / 100)
    supplements = estimate_data.get('supplements', [])
    supplements_total = sum(
        s.get('amount', 0) for s in supplements if s.get('enabled', True)
    )
    tax_amount = (
        (subtotal + op_amount + contingency_amount + supplements_total)
        * (tax_rate / 100) if tax_rate > 0 else 0
    )
    grand_total = (
        subtotal + op_amount + contingency_amount + supplements_total + tax_amount
    )

    # Custom document with footer (page info moved from header to footer)
    class EstimateDocTemplate(BaseDocTemplate):
        def __init__(self, filename, **kwargs):
            self.estimate_number = kwargs.pop('estimate_number', '')
            self.property_address = kwargs.pop('property_address', '')
            self._saved_page_count = 0
            BaseDocTemplate.__init__(self, filename, **kwargs)

            # Define frame for content - more top margin, space for footer
            frame = Frame(
                0.5*inch, 0.6*inch,  # Bottom margin for footer
                letter[0] - 1*inch, letter[1] - 1.1*inch,  # Height adjusted
                id='normal'
            )
            template = PageTemplate(id='estimate', frames=frame, onPage=self.add_page_elements)
            self.addPageTemplates([template])

        def add_page_elements(self, canvas, doc):
            canvas.saveState()

            # ========== FOOTER (bottom of page) ==========
            footer_y = 0.35*inch

            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(colors.Color(0.4, 0.4, 0.4))

            # Page number (bottom left)
            canvas.drawString(0.5*inch, footer_y, f"Page {doc.page}")

            # Estimate number (bottom right)
            canvas.drawRightString(letter[0] - 0.5*inch, footer_y,
                                  f"Estimate #{self.estimate_number}")

            # Property address (bottom center) - on page 2+
            if doc.page > 1 and self.property_address:
                addr_line = self.property_address.replace('\n', ', ')
                if len(addr_line) > 60:
                    addr_line = addr_line[:57] + '...'
                canvas.drawCentredString(letter[0] / 2, footer_y, addr_line)

            canvas.restoreState()

    doc = EstimateDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,  # No header needed now
        bottomMargin=0.6*inch,  # Space for footer
        estimate_number=estimate_number,
        property_address=property_address or '',
    )

    styles = getSampleStyleSheet()

    # Pre-create all styles once (avoid re-instantiation per cell/row)
    style_normal = ParagraphStyle('Normal9', fontSize=9, leading=12)
    style_small = ParagraphStyle('Small', fontSize=9, leading=12, textColor=colors.Color(0.35, 0.35, 0.35))
    style_detail = ParagraphStyle('Detail', fontSize=9, leading=12, textColor=colors.Color(0.45, 0.45, 0.45))
    style_bold = ParagraphStyle('Bold9', fontSize=9, fontName='Helvetica-Bold', leading=12)
    style_section = ParagraphStyle('Section', fontSize=10, fontName='Helvetica-Bold',
                                   textColor=colors.Color(0.2, 0.2, 0.2), spaceBefore=12, spaceAfter=6)
    style_right = ParagraphStyle('Right9', fontSize=9, alignment=TA_RIGHT, leading=12)
    style_right_bold = ParagraphStyle('RB9', fontSize=9, alignment=TA_RIGHT, leading=12, fontName='Helvetica-Bold')
    style_center = ParagraphStyle('C9', fontSize=9, alignment=TA_CENTER, leading=12)
    style_title = ParagraphStyle('Title', fontSize=18, fontName='Helvetica-Bold')
    style_terms = ParagraphStyle('Terms', fontSize=9, leading=12, textColor=colors.Color(0.3, 0.3, 0.3))
    style_total_line = ParagraphStyle('TotalLine', fontSize=14, fontName='Helvetica-Bold')

    story = []

    # ========== HEADER ==========
    company_name = company_info.get('name', 'Moving Company') if company_info else 'Moving Company'
    company_addr = company_info.get('address', '') if company_info else ''
    company_phone = company_info.get('phone', '') if company_info else ''
    company_email = company_info.get('email', '') if company_info else ''

    # Build company info
    company_lines = [f"<b>{company_name}</b>"]
    if company_addr:
        for line in company_addr.split('\n'):
            company_lines.append(line)
    if company_phone:
        company_lines.append(company_phone)
    if company_email:
        company_lines.append(company_email)

    estimate_date = datetime.now().strftime("%B %d, %Y")

    # Header table: Company (left) | Estimate info (right)
    header_left = Paragraph('<br/>'.join(company_lines), style_normal)
    header_right = Paragraph(
        f"""<font size="18"><b>Estimate</b></font><br/><br/>
<font size="9">Estimate number: {estimate_number}<br/>
Estimate date: {estimate_date}</font>""",
        style_right
    )

    header_table = Table([[header_left, header_right]], colWidths=[4*inch, 3.5*inch], hAlign='LEFT')
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),  # Align company info with customer info below
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.25*inch))

    # ========== PREPARED FOR ==========
    story.append(Paragraph("<b>Prepared for</b>", style_bold))
    if client_name:
        story.append(Paragraph(client_name, style_normal))
    if client_phone:
        story.append(Paragraph(client_phone, style_normal))
    if client_email:
        story.append(Paragraph(client_email, style_normal))
    if property_address:
        for line in property_address.split('\n'):
            story.append(Paragraph(line, style_normal))
    else:
        story.append(Paragraph("—", style_normal))
    story.append(Spacer(1, 0.2*inch))

    # ========== TOTAL ESTIMATE ==========
    story.append(Paragraph(f"<b>Total Estimate: ${grand_total:,.2f}</b>", style_total_line))
    story.append(Spacer(1, 0.15*inch))

    # ========== PROJECT DESCRIPTION ==========
    pdf_staging = estimate_data.get('staging_type', 'off_site')
    if pdf_staging == 'on_site':
        story.append(Paragraph("<b>Contents Pack-Out, On-Site Staging & Pack-Back</b>", style_bold))
        description = """Complete contents pack-out of a residential property. Inventory and condition documentation, professional packing and wrapping, furniture disassembly, appliance preparation, relocation to designated on-site staging area, and pack-back with reassembly upon restoration completion."""
    else:
        story.append(Paragraph("<b>Contents Pack-Out, Storage & Pack-Back</b>", style_bold))
        description = """Complete contents pack-out of a water-damaged residential property. Inventory and condition documentation, professional packing and wrapping, furniture disassembly, appliance preparation, transport to climate-controlled storage, and pack-back with reassembly upon restoration completion."""
    story.append(Paragraph(description, style_small))

    # Room breakdown by floor — includes size/density when AI rooms are available,
    # falls back to free-text area_breakdown for quick-estimate path.
    _FLOOR_ORDER = ['basement', '1st', '2nd', '3rd', '4th+']
    _FLOOR_LABEL = {
        'basement': 'Basement', '1st': '1st Floor',
        '2nd': '2nd Floor', '3rd': '3rd Floor', '4th+': '4th Floor+',
    }
    _SIZE_LABEL    = {'small': 'Small', 'large': 'Large', 'xlarge': 'X-Large'}
    _DENSITY_LABEL = {'light': 'Light', 'normal': 'Normal', 'dense': 'Dense',
                      'heavy': 'Heavy', 'extreme': 'Extreme'}

    ai_rooms = estimate_data.get('ai_rooms', [])
    if ai_rooms:
        by_floor: dict = {}
        for r in ai_rooms:
            fl = (r.get('floor') or '1st').lower()
            name = r.get('room_name') or r.get('name', 'Room')
            size = _SIZE_LABEL.get(r.get('room_size') or r.get('roomSize') or 'large', 'Large')
            density = _DENSITY_LABEL.get(r.get('density') or 'normal', 'Normal')
            by_floor.setdefault(fl, []).append(f"{name} ({size} / {density})")
        if by_floor:
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph("<b>Room Breakdown by Floor</b>", style_bold))
            for fl in _FLOOR_ORDER:
                if fl not in by_floor:
                    continue
                label = _FLOOR_LABEL.get(fl, fl.title())
                names = ',  '.join(by_floor[fl])
                story.append(Paragraph(f"<b>{label}:</b>  {names}", style_small))
            story.append(Spacer(1, 0.05*inch))
    elif area_breakdown:
        # Quick-estimate fallback: no AI rooms, show free-text breakdown
        story.append(Spacer(1, 0.1*inch))
        story.append(Paragraph("<b>Area Breakdown</b>", style_bold))
        story.append(Paragraph(
            area_breakdown.replace('\n\n', '<br/><br/>').replace('\n', '<br/>'),
            style_small,
        ))

    story.append(Spacer(1, 0.15*inch))

    # ========== LINE ITEMS BY SECTION ==========
    item_number = 1

    for section in sections:
        # Section header
        story.append(Paragraph(f"<b>{section['title']}</b>", style_section))

        # Table header row — reuse pre-created styles
        table_data = [[
            Paragraph('<b>Item</b>', style_normal),
            Paragraph('<b>Qty</b>', style_right),
            Paragraph('<b>Unit</b>', style_center),
            Paragraph('<b>Price</b>', style_right),
            Paragraph('<b>Total</b>', style_right),
        ]]

        for item in section['items']:
            total = item['qty'] * item['price']

            # Build item cell with number, name, and detail
            if item.get('detail'):
                item_cell = Paragraph(
                    f"<b>{item_number}. {item['name']}</b><br/><font color='#666666'>{item['detail']}</font>",
                    style_normal
                )
            else:
                item_cell = Paragraph(f"<b>{item_number}. {item['name']}</b>", style_normal)

            qty_str = str(int(item['qty'])) if item['qty'] == int(item['qty']) else f"{item['qty']:.1f}"
            table_data.append([
                item_cell,
                Paragraph(qty_str, style_right),
                Paragraph(item['unit'], style_center),
                Paragraph(f"${item['price']:.2f}", style_right),
                Paragraph(f"${total:.2f}", style_right),
            ])
            item_number += 1

        # Section subtotal
        section_subtotal = sum(item['qty'] * item['price'] for item in section['items'])
        table_data.append([
            '', '', '',
            Paragraph('<b>Subtotal:</b>', style_right_bold),
            Paragraph(f"<b>${section_subtotal:,.2f}</b>", style_right_bold),
        ])

        # Create table - full width with proper padding
        # Page width: 8.5", margins: 0.5" each side = 7.5" usable width
        col_widths = [4.5*inch, 0.6*inch, 0.6*inch, 0.8*inch, 1.0*inch]  # Total: 7.5"
        table = Table(table_data, colWidths=col_widths, hAlign='LEFT')
        table.setStyle(TableStyle([
            # Internal padding for all cells
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            # Header row background
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
            ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.Color(0.65, 0.65, 0.65)),
            # Body alignment
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            # Subtle row separators
            ('LINEBELOW', (0, 1), (-1, -2), 0.25, colors.Color(0.85, 0.85, 0.85)),
            # Vertical padding
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.08*inch))

    # ========== GRAND TOTALS ==========
    story.append(Spacer(1, 0.1*inch))

    totals_data = [
        [Paragraph('Items Subtotal', style_right), Paragraph(f"${subtotal:,.2f}", style_right)],
    ]
    if include_op:
        totals_data.append([Paragraph(f'Overhead & Profit ({op_rate}%)', style_right), Paragraph(f"${op_amount:,.2f}", style_right)])
    if include_contingency:
        totals_data.append([
            Paragraph(
                f'Contingency ({contingency_rate}%)', style_right,
            ),
            Paragraph(f"${contingency_amount:,.2f}", style_right),
        ])
    for supp in supplements:
        if supp.get('enabled', True) and supp.get('amount', 0) > 0:
            supp_label = supp['name']
            supp_reason = supp.get('reason', '').strip()
            if supp_reason:
                supp_label += f"<br/><font size=7 color='#666666'>{supp_reason}</font>"
            totals_data.append([
                Paragraph(supp_label, style_right),
                Paragraph(f"${supp['amount']:,.2f}", style_right),
            ])
    if tax_rate > 0:
        totals_data.append([
            Paragraph('Tax', style_right),
            Paragraph(f"${tax_amount:,.2f}", style_right),
        ])
    _total_style = ParagraphStyle(
        'R', fontSize=10,
        fontName='Helvetica-Bold', alignment=TA_RIGHT,
    )
    totals_data.append([
        Paragraph('<b>Total Estimate</b>', _total_style),
        Paragraph(f"<b>${grand_total:,.2f}</b>", _total_style),
    ])

    # Match line item table width: 7.5" total (4.5 + 0.6 + 0.6 + 0.8 + 1.0)
    totals_table = Table(totals_data, colWidths=[6.5*inch, 1.0*inch], hAlign='LEFT')
    totals_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(totals_table)

    # ========== SCHEDULING NOTES ==========
    if notes:
        story.append(Spacer(1, 0.25*inch))
        style_note_hdr = ParagraphStyle(
            'NoteHeader', fontSize=10, leading=13,
            textColor=colors.Color(0.15, 0.15, 0.15),
            spaceAfter=4, fontName='Helvetica-Bold')
        style_note_body = ParagraphStyle(
            'NoteBody', fontSize=9, leading=12,
            textColor=colors.Color(0.25, 0.25, 0.25),
            spaceAfter=4, fontName='Helvetica')
        story.append(Paragraph("Notes", style_note_hdr))
        for line in notes.split("\n"):
            if line.strip():
                story.append(Paragraph(line.strip(), style_note_body))

    # ========== TERMS & CONDITIONS ==========
    story.append(Spacer(1, 0.3*inch))

    style_tc_hdr = ParagraphStyle(
        'TCHeader', fontSize=10, leading=13,
        textColor=colors.Color(0.15, 0.15, 0.15),
        spaceAfter=6, fontName='Helvetica-Bold')
    story.append(Paragraph("Terms &amp; Conditions", style_tc_hdr))
    sm_val = estimate_data.get('storage_months', 1)
    sm_lbl = f"{sm_val} month{'s' if sm_val != 1 else ''}"
    tc_staging = estimate_data.get('staging_type', 'off_site')
    if tc_staging == 'on_site':
        storage_term = (
            "Staging Location",
            "Contents will be staged on-site in a designated area "
            "of the property. The client is responsible for ensuring "
            "the staging area remains accessible and undisturbed "
            "during the restoration period.")
    else:
        storage_term = (
            "Storage Duration",
            f"The base storage period is {sm_lbl}. Additional months "
            "will be billed at the stated monthly rate. The client will "
            "be notified prior to any storage extensions.")
    terms = [
        ("Estimate Validity",
         "This estimate is valid for 30 days from the date of "
         "preparation. Pricing may be adjusted thereafter based on "
         "current market conditions."),
        ("Scope Changes",
         "Any work beyond the scope described in this estimate will "
         "require a written Change Order approved by the client or "
         "their representative prior to commencement."),
        storage_term,
        ("Liability",
         "We are not responsible for pre-existing damage, mold, or "
         "deterioration caused by water damage prior to our arrival. "
         "All items are documented and photographed at the time of "
         "pack-out."),
        ("Access &amp; Scheduling",
         "The client is responsible for providing clear access to the "
         "property for pack-out and pack-back. Delays caused by "
         "restricted access or scheduling conflicts may result in "
         "additional charges."),
        ("Exclusions",
         "This estimate does not include textile/fabric restoration "
         "cleaning, content replacement or repair, mold remediation, "
         "structural drying, or any third-party specialty services."),
    ]
    if estimate_data.get('include_insurance_clauses', False):
        terms.append((
            "Insurance Coordination",
            "This estimate is prepared for submission to the property "
            "insurance carrier. Final payment terms are subject to "
            "the carrier\u2019s approval and the client\u2019s policy coverage."))
        terms.append((
            "Payment",
            "Payment is due upon receipt of invoice. For insurance "
            "claims, payment is expected within 7 business days of "
            "receipt from the carrier."))
    for i, (title, body) in enumerate(terms, 1):
        story.append(Paragraph(
            f"{i}. <b>{title}:</b> {body}", style_terms))
        story.append(Spacer(1, 4))

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================
# EXCEL EXPORT
# ============================================

def generate_estimate_excel(
    estimate_data: Dict[str, Any],
    client_name: Optional[str] = None,
    client_phone: Optional[str] = None,
    client_email: Optional[str] = None,
    property_address: Optional[str] = None,
    notes: Optional[str] = None,
    company_info: Optional[Dict] = None,
    estimate_number: Optional[str] = None,
    tax_rate: float = 0,
    prices: Optional[Dict[str, float]] = None,
) -> bytes:
    """Generate professional Excel estimate"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=9)
    section_font = Font(bold=True, size=10)
    normal_font = Font(size=9)
    detail_font = Font(size=8, color='666666')
    total_font = Font(bold=True, size=11)

    header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    total_fill = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid')
    thin_border = Border(bottom=Side(style='thin', color='D9D9D9'))

    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    if not estimate_number:
        estimate_number = generate_estimate_number()

    row = 1

    # Company info
    if company_info:
        ws[f'A{row}'] = company_info.get('name', 'Moving Company')
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        if company_info.get('address'):
            for line in company_info.get('address', '').split('\n'):
                ws[f'A{row}'] = line
                row += 1
        if company_info.get('phone'):
            ws[f'A{row}'] = company_info.get('phone')
            row += 1
        if company_info.get('email'):
            ws[f'A{row}'] = company_info.get('email')
            row += 1
        row += 1

    # Estimate header
    ws[f'A{row}'] = 'Estimate'
    ws[f'A{row}'].font = title_font
    row += 1
    ws[f'A{row}'] = f'Estimate number: {estimate_number}'
    row += 1
    ws[f'A{row}'] = f'Estimate date: {datetime.now().strftime("%B %d, %Y")}'
    row += 2

    # Client info
    if client_name:
        ws[f'A{row}'] = 'Prepared for'
        ws[f'A{row}'].font = Font(bold=True, size=9)
        row += 1
        ws[f'A{row}'] = client_name
        row += 1
    if client_phone:
        ws[f'A{row}'] = client_phone
        row += 1
    if client_email:
        ws[f'A{row}'] = client_email
        row += 1
    if property_address:
        for line in property_address.split('\n'):
            ws[f'A{row}'] = line
            row += 1
    row += 1

    # Priority: 1) frontend line_items (editor edits), 2) section_details from
    # calculate_estimate_from_content, 3) recalculate from totals (legacy fallback).
    # Using section_details ensures PDF/Excel matches the Estimate Editor exactly.
    line_items = estimate_data.get('line_items') or []
    section_details = estimate_data.get('section_details') or {}
    if line_items:
        sections = line_items
    elif section_details:
        sections = _section_details_to_line_items(
            section_details,
            estimate_data.get('sections', {}),
            estimate_data.get('material_details'),
        )
    else:
        sections = get_line_items_from_estimate(estimate_data, prices=prices)

    include_op = estimate_data.get('include_op', False)
    op_rate = estimate_data.get('op_rate', 10)
    include_contingency = estimate_data.get('include_contingency', False)
    contingency_rate = estimate_data.get('contingency_rate', 5)
    subtotal = sum(
        item['qty'] * item['price']
        for section in sections
        for item in section['items']
    )
    # Respect the include flags — if disabled, force 0 regardless of saved value
    op_amount = 0
    if include_op:
        saved_op = estimate_data.get('op_amount')
        op_amount = saved_op if saved_op is not None else subtotal * (op_rate / 100)
    contingency_amount = 0
    if include_contingency:
        saved_cont = estimate_data.get('contingency_amount')
        contingency_amount = saved_cont if saved_cont is not None else subtotal * (contingency_rate / 100)
    supplements = estimate_data.get('supplements', [])
    supplements_total = sum(
        s.get('amount', 0) for s in supplements if s.get('enabled', True)
    )
    tax_amount = (
        (subtotal + op_amount + contingency_amount + supplements_total)
        * (tax_rate / 100) if tax_rate > 0 else 0
    )
    grand_total = (
        subtotal + op_amount + contingency_amount + supplements_total + tax_amount
    )

    # Total estimate box
    ws[f'A{row}'] = f'Total Estimate: ${grand_total:,.2f}'
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'A{row}'].fill = total_fill
    row += 2

    # Line items by section
    item_number = 1

    for section in sections:
        ws[f'A{row}'] = section['title']
        ws[f'A{row}'].font = section_font
        row += 1

        # Column headers
        headers = ['Item', 'Qty', 'Unit', 'Price', 'Total']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        section_start = row
        for item in section['items']:
            ws.cell(row=row, column=1, value=f"{item_number}. {item['name']}").font = normal_font
            ws.cell(row=row, column=2, value=item['qty']).font = normal_font
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=3, value=item['unit']).font = normal_font
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=item['price']).font = normal_font
            ws.cell(row=row, column=4).number_format = '$#,##0.00'
            ws.cell(row=row, column=5, value=f'=B{row}*D{row}').font = normal_font
            ws.cell(row=row, column=5).number_format = '$#,##0.00'

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border

            item_number += 1
            row += 1

        # Section subtotal
        ws.cell(row=row, column=4, value='Subtotal:').font = Font(bold=True, size=9)
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=5, value=f'=SUM(E{section_start}:E{row-1})').font = Font(bold=True, size=9)
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        row += 2

    # Grand totals
    ws.cell(row=row, column=4, value='Items Subtotal').font = normal_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=5, value=subtotal).font = normal_font
    ws.cell(row=row, column=5).number_format = '$#,##0.00'
    row += 1

    if include_op:
        ws.cell(row=row, column=4, value=f'Overhead & Profit ({op_rate}%)').font = normal_font
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=5, value=op_amount).font = normal_font
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        row += 1

    if include_contingency:
        ws.cell(
            row=row, column=4,
            value=f'Contingency ({contingency_rate}%)',
        ).font = normal_font
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=5, value=contingency_amount).font = normal_font
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        row += 1

    for supp in supplements:
        if supp.get('enabled', True) and supp.get('amount', 0) > 0:
            supp_label = supp['name']
            supp_reason = supp.get('reason', '').strip()
            if supp_reason:
                supp_label += f" ({supp_reason})"
            ws.cell(
                row=row, column=4, value=supp_label,
            ).font = normal_font
            ws.cell(row=row, column=4).alignment = Alignment(
                horizontal='right',
            )
            ws.cell(
                row=row, column=5, value=supp['amount'],
            ).font = normal_font
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            row += 1

    if tax_rate > 0:
        ws.cell(row=row, column=4, value='Tax').font = normal_font
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=5, value=tax_amount).font = normal_font
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        row += 1

    ws.cell(row=row, column=4, value='Total Estimate').font = total_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=5, value=grand_total).font = total_font
    ws.cell(row=row, column=5).number_format = '$#,##0.00'
    row += 2

    # ========== TERMS & CONDITIONS ==========
    tc_header_font = Font(bold=True, size=10)
    tc_font = Font(size=8, color='333333')
    ws.cell(row=row, column=1, value='Terms & Conditions').font = tc_header_font
    row += 1

    sm_val = estimate_data.get('storage_months', 1)
    sm_lbl = f"{sm_val} month{'s' if sm_val != 1 else ''}"
    xl_staging = estimate_data.get('staging_type', 'off_site')
    if xl_staging == 'on_site':
        xl_storage_term = (
            "Staging Location",
            "Contents will be staged on-site in a designated area "
            "of the property. The client is responsible for ensuring "
            "the staging area remains accessible and undisturbed "
            "during the restoration period.")
    else:
        xl_storage_term = (
            "Storage Duration",
            f"The base storage period is {sm_lbl}. Additional months "
            "will be billed at the stated monthly rate. The client "
            "will be notified prior to any storage extensions.")
    tc_terms = [
        ("Estimate Validity",
         "This estimate is valid for 30 days from the date of "
         "preparation. Pricing may be adjusted thereafter based "
         "on current market conditions."),
        ("Scope Changes",
         "Any work beyond the scope described in this estimate "
         "will require a written Change Order approved by the "
         "client or their representative prior to commencement."),
        xl_storage_term,
        ("Liability",
         "We are not responsible for pre-existing damage, mold, or "
         "deterioration caused by water damage prior to our arrival. "
         "All items are documented and photographed at the time of "
         "pack-out."),
        ("Access & Scheduling",
         "The client is responsible for providing clear access to "
         "the property for pack-out and pack-back. Delays caused by "
         "restricted access or scheduling conflicts may result in "
         "additional charges."),
        ("Exclusions",
         "This estimate does not include textile/fabric restoration "
         "cleaning, content replacement or repair, mold remediation, "
         "structural drying, or any third-party specialty services."),
    ]
    if estimate_data.get('include_insurance_clauses', False):
        tc_terms.append((
            "Insurance Coordination",
            "This estimate is prepared for submission to the property "
            "insurance carrier. Final payment terms are subject to "
            "the carrier\u2019s approval and the client\u2019s policy "
            "coverage."))
        tc_terms.append((
            "Payment",
            "Payment is due upon receipt of invoice. For insurance "
            "claims, payment is expected within 7 business days of "
            "receipt from the carrier."))
    for i, (title, body) in enumerate(tc_terms, 1):
        cell = ws.cell(row=row, column=1,
                       value=f"{i}. {title}: {body}")
        cell.font = tc_font
        cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=5)
        row += 1

    # ── Notes (e.g. workday scheduling) ──
    if notes:
        row += 1
        cell = ws.cell(row=row, column=1, value="Notes")
        cell.font = Font(bold=True, size=10)
        row += 1
        for line in notes.split("\n"):
            if line.strip():
                cell = ws.cell(row=row, column=1, value=line.strip())
                cell.font = Font(size=9, color="444444")
                cell.alignment = Alignment(wrap_text=True)
                ws.merge_cells(start_row=row, start_column=1,
                               end_row=row, end_column=5)
                row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================
# PACK-OUT INVENTORY EXCEL EXPORT
# ============================================

def generate_inventory_excel(
    rooms: List[Dict[str, Any]],
    property_address: str = "",
    claim_number: str = "",
    prepared_by: str = "",
) -> bytes:
    """Generate a professional Pack-Out Inventory Excel
    matching the 13-column reference format.

    ``rooms`` is a list of dicts, each with:
      - room_name: str
      - items: list of item dicts from RoomAnalysisResponse
      - field_notes: optional list of str
      - total_labor_hours: optional float
      - fragile_count: optional int
      - high_value_count: optional int
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Pack-Out Inventory"

    # --- Styles ---
    title_font = Font(
        name="Calibri", size=14, bold=True, color="1F4E79"
    )
    subtitle_font = Font(
        name="Calibri", size=10, color="555555"
    )
    header_font = Font(
        name="Calibri", size=10, bold=True, color="FFFFFF"
    )
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79",
        fill_type="solid"
    )
    room_font = Font(
        name="Calibri", size=11, bold=True, color="1F4E79"
    )
    room_fill = PatternFill(
        start_color="D6E4F0", end_color="D6E4F0",
        fill_type="solid"
    )
    data_font = Font(name="Calibri", size=9)
    flag_font = Font(
        name="Calibri", size=9, bold=True, color="CC0000"
    )
    summary_font = Font(
        name="Calibri", size=10, bold=True
    )
    note_font = Font(
        name="Calibri", size=9, italic=True
    )
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    wrap_align = Alignment(
        vertical="top", wrap_text=True
    )

    # --- Column widths ---
    col_widths = {
        1: 9, 2: 18, 3: 40, 4: 18, 5: 5,
        6: 8, 7: 40, 8: 24, 9: 10,
        10: 35, 11: 10, 12: 16,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[
            get_column_letter(col)
        ].width = w

    # --- Title rows ---
    today = datetime.now().strftime("%Y-%m-%d")
    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1,
                value="CONTENT PACK-OUT INVENTORY & WORK SCOPE")
    c.font = title_font

    sub = (
        f"Property: {property_address or '___'}  |  "
        f"Date: {today}  |  "
        f"Claim #: {claim_number or '___'}  |  "
        f"Prepared by: {prepared_by or '___'}"
    )
    ws.merge_cells("A2:L2")
    c2 = ws.cell(row=2, column=1, value=sub)
    c2.font = subtitle_font

    # --- Column headers (row 4) ---
    headers = [
        "Item #", "Location / Room", "Item Description",
        "Category", "Qty", "Fragile",
        "Packing Method", "Box Size / Material",
        "Est. Labor", "Notes / Special Instructions",
        "Photo Ref", "Estimator Flag",
    ]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center",
            wrap_text=True,
        )
        cell.border = thin_border

    row = 5
    grand_labor = 0.0
    grand_fragile = 0
    grand_hv = 0
    all_field_notes: List[str] = []

    # --- Room sections ---
    for room in rooms:
        room_name = room.get("room_name", "Unknown Room")

        # Room separator row
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=12,
        )
        rc = ws.cell(row=row, column=1, value=room_name)
        rc.font = room_font
        rc.fill = room_fill
        rc.border = thin_border
        row += 1

        items = room.get("items", [])
        room_prefix = _room_prefix(room_name)

        for idx, item in enumerate(items, 1):
            item_num = f"{room_prefix}-{idx:03d}"
            qty = item.get("quantity", 1)
            fragile = "Y" if item.get("is_fragile") else "N"
            _bh = item.get("base_labor_hours")
            _puh = item.get("per_unit_labor_hours")
            if _bh is not None and _puh is not None:
                labor = _bh + (_puh * qty)
            else:
                labor = item.get("estimated_labor_hours")
            labor_str = (
                f"{labor:.2f}" if labor is not None else ""
            )
            if labor:
                grand_labor += labor  # already total

            flags = item.get("estimator_flags") or []
            flag_str = ", ".join(flags) if flags else ""

            if item.get("is_fragile"):
                grand_fragile += qty
            if item.get("is_high_value"):
                grand_hv += qty

            mats = item.get("required_materials") or []
            mats_str = ", ".join(
                m.replace("_", " ").title() for m in mats
            )

            values = [
                item_num,
                room_name,
                item.get("name", ""),
                item.get("category", ""),
                qty,
                fragile,
                item.get("packing_method", ""),
                mats_str,
                labor_str,
                item.get("special_instructions", "") or "",
                "",  # photo ref — not tracked yet
                flag_str,
            ]

            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = data_font
                cell.alignment = wrap_align
                cell.border = thin_border
                if ci == 12 and flag_str:
                    cell.font = flag_font

            row += 1

        # Collect field notes from room
        for note in room.get("field_notes", []):
            all_field_notes.append(f"[{room_name}] {note}")

    # --- Summary section ---
    row += 1
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=12,
    )
    ws.cell(
        row=row, column=1, value="SUMMARY & TOTALS"
    ).font = Font(
        name="Calibri", size=12, bold=True,
        color="1F4E79",
    )
    row += 1

    summary_rows = [
        ("Total Estimated Labor Hours:", f"{grand_labor:.1f}",
         "hrs"),
        ("Total Fragile Items:", str(grand_fragile),
         "items"),
        ("Total High-Value Items:", str(grand_hv),
         "items"),
    ]
    for label, val, unit in summary_rows:
        ws.cell(row=row, column=1, value=label).font = (
            summary_font
        )
        ws.cell(row=row, column=9, value=val).font = (
            summary_font
        )
        ws.cell(row=row, column=10, value=unit).font = (
            data_font
        )
        row += 1

    # --- Field Notes (regenerated from current items at report time) ---
    if all_field_notes:
        row += 1
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=12,
        )
        ws.cell(
            row=row, column=1,
            value="FIELD NOTES & ESTIMATOR OBSERVATIONS",
        ).font = Font(
            name="Calibri", size=12, bold=True,
            color="1F4E79",
        )
        row += 1

        for i, note in enumerate(all_field_notes[:15], 1):
            ws.merge_cells(
                start_row=row, start_column=1,
                end_row=row, end_column=12,
            )
            cell = ws.cell(
                row=row, column=1,
                value=f"{i}. {note}",
            )
            cell.font = note_font
            cell.alignment = wrap_align
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _room_prefix(room_name: str) -> str:
    """Generate a short prefix from room name for item IDs.
    e.g. '2nd Floor - Room 1' -> 'R1',
         'Bathroom' -> 'BA'
    """
    name = room_name.upper().strip()
    # Try to extract room number
    m = re.search(r'ROOM\s*(\d+)', name)
    if m:
        return f"R{m.group(1)}"
    m = re.search(r'OFFICE\s*(\d+)', name)
    if m:
        return f"O{m.group(1)}"

    abbrevs = {
        "BEDROOM": "BD", "KITCHEN": "KT", "LIVING": "LV",
        "DINING": "DN", "BATHROOM": "BA", "CLOSET": "CL",
        "HALLWAY": "HL", "OFFICE": "OF", "GARAGE": "GA",
        "BASEMENT": "BS", "LAUNDRY": "LA", "ATTIC": "AT",
    }
    for keyword, abbr in abbrevs.items():
        if keyword in name:
            return abbr
    # Fallback: first 2 chars
    clean = re.sub(r'[^A-Z]', '', name)
    return clean[:2] or "RM"


# ============================================
# REPORT PDF EXPORT
# ============================================

def _compress_image_bytes(
    img_data: bytes,
    max_width: int = 800,
    max_height: int = 800,
    quality: int = 60,
) -> bytes:
    """Compress an image to JPEG with size limits for small PDF output."""
    from PIL import Image, ImageOps
    img = Image.open(io.BytesIO(img_data))

    # Fix EXIF orientation (iPhone photos are often rotated in EXIF metadata)
    try:
        img = ImageOps.exif_transpose(img)
    except Exception:
        pass  # If EXIF data is missing or corrupt, continue with original

    # Convert RGBA/P to RGB for JPEG
    if img.mode in ("RGBA", "P"):
        bg = Image.new("RGB", img.size, (255, 255, 255))
        if img.mode == "P":
            img = img.convert("RGBA")
        bg.paste(img, mask=img.split()[3] if img.mode == "RGBA" else None)
        img = bg
    elif img.mode != "RGB":
        img = img.convert("RGB")

    # Resize if width OR height exceeds limits
    if img.width > max_width or img.height > max_height:
        ratio = min(max_width / img.width, max_height / img.height)
        new_w = max(1, int(img.width * ratio))
        new_h = max(1, int(img.height * ratio))
        img = img.resize((new_w, new_h), Image.LANCZOS)

    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=quality, optimize=True)
    return buf.getvalue()


def _b64_to_compressed_image(
    b64_str: str,
    max_width: int = 800,
    max_height: int = 800,
    quality: int = 60,
) -> bytes:
    """Decode base64 image and compress it."""
    import base64
    # Strip data URI prefix if present
    if "," in b64_str:
        b64_str = b64_str.split(",", 1)[1]
    raw = base64.b64decode(b64_str)
    return _compress_image_bytes(
        raw, max_width=max_width, max_height=max_height,
        quality=quality,
    )


def generate_report_pdf(
    session_data: Dict[str, Any],
    rooms_data: List[Dict[str, Any]],
    sections_config: Dict[str, bool],
    client_name: Optional[str] = None,
    client_phone: Optional[str] = None,
    client_email: Optional[str] = None,
    property_address: Optional[str] = None,
    company_info: Optional[Dict] = None,
    tax_rate: float = 0,
    notes: Optional[str] = None,
    include_signature_page: bool = False,
    image_quality: int = 60,
    max_image_width: int = 800,
) -> bytes:
    """Generate a clean, professional packing report PDF.

    Sections are toggled via ``sections_config``:
      - inventory_list: room-by-room item inventory
      - damage_photos: pre-existing damage photos
      - labor_log: labor hours per room
      - room_photos: general room photos
      - estimate_summary: estimate totals
    """
    buffer = io.BytesIO()
    estimate_data = session_data.get("result") or session_data
    settings = session_data.get("settings", {})
    estimate_data = {**settings, **estimate_data}

    report_number = generate_estimate_number().replace("EST-", "RPT-")
    report_date = datetime.now().strftime("%B %d, %Y")

    # --- Document setup ---
    class ReportDocTemplate(BaseDocTemplate):
        def __init__(self, filename, **kwargs):
            self.report_number = kwargs.pop("report_number", "")
            self.property_address = kwargs.pop("property_address", "")
            BaseDocTemplate.__init__(self, filename, **kwargs)
            frame = Frame(
                0.5 * inch, 0.6 * inch,
                letter[0] - 1 * inch, letter[1] - 1.1 * inch,
                id="normal",
            )
            template = PageTemplate(
                id="report", frames=frame, onPage=self._page_elements,
            )
            self.addPageTemplates([template])

        def _page_elements(self, cvs, doc):
            cvs.saveState()
            footer_y = 0.35 * inch
            cvs.setFont("Helvetica", 8)
            cvs.setFillColor(colors.Color(0.4, 0.4, 0.4))
            cvs.drawString(0.5 * inch, footer_y, f"Page {doc.page}")
            cvs.drawRightString(
                letter[0] - 0.5 * inch, footer_y,
                f"Report #{self.report_number}",
            )
            if doc.page > 1 and self.property_address:
                addr = self.property_address.replace("\n", ", ")
                if len(addr) > 60:
                    addr = addr[:57] + "..."
                cvs.drawCentredString(letter[0] / 2, footer_y, addr)
            cvs.restoreState()

    doc = ReportDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.6 * inch,
        report_number=report_number,
        property_address=property_address or "",
    )

    # --- Styles ---
    style_normal = ParagraphStyle("RN", fontSize=9, leading=12)
    style_small = ParagraphStyle(
        "RS", fontSize=8, leading=10,
        textColor=colors.Color(0.4, 0.4, 0.4),
    )
    style_bold = ParagraphStyle(
        "RB", fontSize=9, fontName="Helvetica-Bold", leading=12,
    )
    style_section = ParagraphStyle(
        "RSec", fontSize=11, fontName="Helvetica-Bold",
        textColor=colors.Color(0.15, 0.15, 0.15),
        spaceBefore=14, spaceAfter=6,
    )
    style_room_header = ParagraphStyle(
        "RRoom", fontSize=10, fontName="Helvetica-Bold",
        textColor=colors.Color(0.2, 0.2, 0.2),
        spaceBefore=10, spaceAfter=4,
    )
    style_right = ParagraphStyle(
        "RR", fontSize=9, alignment=TA_RIGHT, leading=12,
    )
    story = []

    # ========== HEADER ==========
    co_name = company_info.get("name", "Company") if company_info else "Company"
    co_addr = company_info.get("address", "") if company_info else ""
    co_phone = company_info.get("phone", "") if company_info else ""
    co_email = company_info.get("email", "") if company_info else ""

    co_lines = [f"<b>{co_name}</b>"]
    if co_addr:
        for line in co_addr.split("\n"):
            co_lines.append(line)
    if co_phone:
        co_lines.append(co_phone)
    if co_email:
        co_lines.append(co_email)

    header_left = Paragraph("<br/>".join(co_lines), style_normal)
    header_right = Paragraph(
        f'<font size="16"><b>Packing Report</b></font><br/><br/>'
        f'<font size="9">Report #: {report_number}<br/>'
        f"Date: {report_date}</font>",
        style_right,
    )
    header_table = Table(
        [[header_left, header_right]],
        colWidths=[4 * inch, 3.5 * inch],
        hAlign="LEFT",
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (0, 0), 0),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.2 * inch))

    # ========== CLIENT INFO ==========
    story.append(Paragraph("<b>Prepared for</b>", style_bold))
    if client_name:
        story.append(Paragraph(client_name, style_normal))
    if client_phone:
        story.append(Paragraph(client_phone, style_normal))
    if client_email:
        story.append(Paragraph(client_email, style_normal))
    if property_address:
        for line in property_address.split("\n"):
            story.append(Paragraph(line, style_normal))
    story.append(Spacer(1, 0.15 * inch))

    # Divider
    story.append(HRFlowable(
        width="100%", thickness=0.5,
        color=colors.Color(0.8, 0.8, 0.8),
    ))
    story.append(Spacer(1, 0.1 * inch))

    # ========== ESTIMATE SUMMARY ==========
    show_summary = sections_config.get("estimate_summary", True)
    if show_summary:
        grand_total = estimate_data.get("grand_total", 0)
        subtotal = estimate_data.get("subtotal", 0)
        total_rooms = estimate_data.get("total_rooms", 0)
        total_hours = estimate_data.get("total_hours", 0)
        crew_size = estimate_data.get("crew_size", 4)

        # Cross-reference: link to the estimate that generated this report
        estimate_id = estimate_data.get("id") or session_data.get("id") or ""
        story.append(Paragraph("Estimate Summary", style_section))
        if estimate_id:
            story.append(Paragraph(
                f'<font size="8" color="#666666">Ref: Estimate #{estimate_id}  |  Report #{report_number}</font>',
                style_small,
            ))
            story.append(Spacer(1, 4))

        summary_data = [
            ["Rooms", "Hours", "Crew", "Subtotal", "Grand Total"],
            [
                str(total_rooms),
                str(round(total_hours, 1)),
                str(crew_size),
                f"${subtotal:,.2f}",
                f"${grand_total:,.2f}",
            ],
        ]
        summary_table = Table(
            summary_data,
            colWidths=[1.1 * inch] * 5,
            hAlign="LEFT",
        )
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.93, 0.93, 0.93)),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0.85, 0.85, 0.85)),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.15 * inch))

    # ========== ROOM-BY-ROOM SECTIONS ==========
    show_inventory = sections_config.get("inventory_list", True)
    show_damage = sections_config.get("damage_photos", False)
    show_labor = sections_config.get("labor_log", False)
    show_photos = sections_config.get("room_photos", True)

    if rooms_data and (show_inventory or show_damage or show_photos or show_labor):
        for room_idx, room in enumerate(rooms_data):
            room_name = room.get("room_name", f"Room {room_idx + 1}")

            story.append(Paragraph(room_name, style_room_header))
            story.append(HRFlowable(
                width="100%", thickness=0.3,
                color=colors.Color(0.85, 0.85, 0.85),
            ))
            story.append(Spacer(1, 4))

            # -- Inventory List --
            items = room.get("items") or []
            if show_inventory and items:
                inv_header = ["#", "Item", "Category", "Qty", "Notes"]
                inv_rows = [inv_header]
                for i, item in enumerate(items, 1):
                    flags = []
                    if item.get("is_fragile"):
                        flags.append("Fragile")
                    if item.get("is_high_value"):
                        flags.append("High Value")
                    if item.get("needs_disassembly"):
                        flags.append("Disassembly")
                    notes_str = ", ".join(flags) if flags else ""
                    inv_rows.append([
                        str(i),
                        Paragraph(item.get("name", ""), style_normal),
                        item.get("category", ""),
                        str(item.get("quantity", 1)),
                        Paragraph(notes_str, style_small),
                    ])

                inv_table = Table(
                    inv_rows,
                    colWidths=[0.35 * inch, 2.5 * inch, 1.3 * inch, 0.5 * inch, 2.35 * inch],
                    hAlign="LEFT",
                    repeatRows=1,
                )
                inv_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.95, 0.95, 0.95)),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("ALIGN", (3, 0), (3, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.Color(0.88, 0.88, 0.88)),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]))
                story.append(KeepTogether([
                    Paragraph(f"<b>Inventory ({len(items)} items)</b>", style_bold),
                    Spacer(1, 3),
                    inv_table,
                ]))
                story.append(Spacer(1, 8))

            # -- Labor Log --
            # Always show labor info when items exist (even if labor_log toggle is off),
            # as a compact summary line under the inventory. Full labor table only
            # when labor_log is explicitly enabled.
            labor_hours = room.get("labor_hours") or 0
            labor_notes = room.get("labor_notes") or ""
            if items and (labor_hours > 0 or labor_notes):
                if show_labor:
                    # Full labor table
                    labor_data = [
                        ["Labor Hours", "Notes"],
                        [
                            f"{labor_hours:.1f} hrs",
                            Paragraph(labor_notes, style_small),
                        ],
                    ]
                    labor_table = Table(
                        labor_data,
                        colWidths=[1.2 * inch, 5.8 * inch],
                        hAlign="LEFT",
                    )
                    labor_table.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.95, 0.95, 0.95)),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.Color(0.88, 0.88, 0.88)),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]))
                    story.append(KeepTogether([
                        Paragraph("<b>Labor</b>", style_bold),
                        Spacer(1, 3),
                        labor_table,
                    ]))
                    story.append(Spacer(1, 8))
                else:
                    # Compact summary line (always shown when items exist)
                    summary_text = f"Labor: {labor_hours:.1f} hrs"
                    if labor_notes:
                        summary_text += f" — {labor_notes}"
                    story.append(Paragraph(summary_text, style_small))
                    story.append(Spacer(1, 4))

            # -- Room Photos & Damage Photos --
            photos = room.get("photos") or []
            general_photos = [p for p in photos if not p.get("is_damage")]
            damage_photos = [p for p in photos if p.get("is_damage")]

            if show_photos and general_photos:
                story.append(Paragraph("<b>Room Photos</b>", style_bold))
                story.append(Spacer(1, 3))
                _add_photo_grid(
                    story, general_photos, style_small,
                    max_width=max_image_width, quality=image_quality,
                )
                story.append(Spacer(1, 8))

            if show_damage and damage_photos:
                story.append(Paragraph(
                    "<b>Pre-Existing Damage</b>", style_bold,
                ))
                story.append(Spacer(1, 3))
                _add_photo_grid(
                    story, damage_photos, style_small,
                    max_width=max_image_width, quality=image_quality,
                )
                story.append(Spacer(1, 8))

            # Field notes (regenerated from current items at report time)
            field_notes = room.get("field_notes") or []
            if field_notes:
                notes_text = " | ".join(field_notes)
                story.append(Paragraph(
                    f"<i>Notes: {notes_text}</i>", style_small,
                ))
                story.append(Spacer(1, 6))

    # ========== INVENTORY SUMMARY ==========
    if rooms_data and show_inventory:
        # Aggregate stats across all rooms
        total_line_items = 0
        total_qty = 0
        cat_counts: Dict[str, int] = {}
        fragile_total = 0
        hv_total = 0
        heavy_total = 0
        disassembly_total = 0

        for room in rooms_data:
            items = room.get("items") or []
            total_line_items += len(items)
            for item in items:
                qty = item.get("quantity", 1) or 1
                total_qty += qty
                cat = item.get("category", "Other")
                cat_counts[cat] = cat_counts.get(cat, 0) + qty
                if item.get("is_fragile"):
                    fragile_total += qty
                if item.get("is_high_value"):
                    hv_total += qty
                if item.get("weight") in ("heavy", "extra_heavy"):
                    heavy_total += qty
                if item.get("needs_disassembly"):
                    disassembly_total += qty

        story.append(Spacer(1, 0.1 * inch))
        story.append(HRFlowable(
            width="100%", thickness=0.5,
            color=colors.Color(0.8, 0.8, 0.8),
        ))
        story.append(Spacer(1, 6))
        story.append(Paragraph("Inventory Summary", style_section))

        # Top-level stats
        summary_rows = [
            ["Total Rooms", "Line Items", "Total Pieces", "Fragile", "High-Value", "Heavy"],
            [
                str(len(rooms_data)),
                str(total_line_items),
                str(total_qty),
                str(fragile_total),
                str(hv_total),
                str(heavy_total),
            ],
        ]
        sum_table = Table(
            summary_rows,
            colWidths=[1.17 * inch] * 6,
            hAlign="LEFT",
        )
        sum_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.93, 0.93, 0.93)),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0.85, 0.85, 0.85)),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(sum_table)

        # Category breakdown (top 8 categories)
        if cat_counts:
            sorted_cats = sorted(cat_counts.items(), key=lambda x: -x[1])[:8]
            cat_header = [c[0] for c in sorted_cats]
            cat_values = [str(c[1]) for c in sorted_cats]
            cat_table = Table(
                [cat_header, cat_values],
                colWidths=[max(0.8, 7.0 / len(sorted_cats)) * inch] * len(sorted_cats),
                hAlign="LEFT",
            )
            cat_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.95, 0.95, 0.95)),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.Color(0.88, 0.88, 0.88)),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]))
            story.append(Spacer(1, 6))
            story.append(Paragraph("<b>Category Breakdown</b>", style_bold))
            story.append(Spacer(1, 3))
            story.append(cat_table)

        story.append(Spacer(1, 0.15 * inch))

    # ========== ADDITIONAL NOTES ==========
    if notes:
        story.append(Spacer(1, 0.1 * inch))
        story.append(HRFlowable(
            width="100%", thickness=0.5,
            color=colors.Color(0.8, 0.8, 0.8),
        ))
        story.append(Spacer(1, 6))
        story.append(Paragraph("<b>Additional Notes</b>", style_bold))
        story.append(Spacer(1, 3))
        story.append(Paragraph(notes, style_normal))

    # ========== SIGNATURE PAGE ==========
    if include_signature_page:
        story.append(PageBreak())
        story.append(Spacer(1, 0.5 * inch))
        story.append(Paragraph(
            "<b>Acknowledgment & Authorization</b>",
            ParagraphStyle(
                "SigTitle", fontSize=14, fontName="Helvetica-Bold",
                alignment=TA_CENTER, spaceBefore=20, spaceAfter=10,
            ),
        ))
        story.append(Spacer(1, 0.2 * inch))

        ack_text = (
            "I acknowledge that I have reviewed this packing report and the "
            "inventory of contents listed herein. I understand that this report "
            "documents the condition and contents of the property at the time "
            "of the pack-out. I authorize the listed company to proceed with "
            "the packing, transport, and storage of the contents as described."
        )
        story.append(Paragraph(ack_text, style_normal))
        story.append(Spacer(1, 0.6 * inch))

        # Signature lines
        sig_data = [
            [
                Paragraph("<b>Customer Signature</b>", style_bold),
                "",
                Paragraph("<b>Date</b>", style_bold),
            ],
            [
                "",
                "",
                "",
            ],
            [
                "________________________________________",
                "",
                "____________________",
            ],
        ]
        sig_table = Table(
            sig_data,
            colWidths=[3.5 * inch, 0.5 * inch, 2.5 * inch],
            hAlign="LEFT",
        )
        sig_table.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ]))
        story.append(sig_table)
        story.append(Spacer(1, 0.5 * inch))

        # Company rep signature
        rep_data = [
            [
                Paragraph("<b>Company Representative</b>", style_bold),
                "",
                Paragraph("<b>Date</b>", style_bold),
            ],
            ["", "", ""],
            [
                "________________________________________",
                "",
                "____________________",
            ],
        ]
        rep_table = Table(
            rep_data,
            colWidths=[3.5 * inch, 0.5 * inch, 2.5 * inch],
            hAlign="LEFT",
        )
        rep_table.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ]))
        story.append(rep_table)

        if client_name:
            story.append(Spacer(1, 0.3 * inch))
            story.append(Paragraph(
                f"<b>Customer:</b> {client_name}", style_normal,
            ))
        if property_address:
            story.append(Paragraph(
                f"<b>Property:</b> {property_address}", style_normal,
            ))

    # Build
    doc.build(story)
    return buffer.getvalue()


def _add_photo_grid(
    story: list,
    photos: List[Dict],
    caption_style,
    max_width: int = 800,
    quality: int = 60,
    cols: int = 2,
):
    """Add photos in a grid layout to the story.

    Each photo dict has 'image' (base64) and optional 'caption'.
    Photos are compressed before embedding.
    """
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.utils import ImageReader

    cell_width = (7.0 * inch) / cols
    img_display_w = cell_width - 0.2 * inch
    max_img_h = 2.2 * inch

    row_cells = []
    rows = []

    for photo in photos:
        b64 = photo.get("image", "")
        if not b64 or len(b64) < 100:
            continue
        try:
            compressed = _b64_to_compressed_image(
                b64, max_width=max_width, quality=quality,
            )
            img_reader = ImageReader(io.BytesIO(compressed))
            iw, ih = img_reader.getSize()

            # Guard against corrupt / zero dimensions
            if iw <= 0 or ih <= 0 or iw > 20000 or ih > 20000:
                continue

            # Scale to fit cell with hard caps
            ratio = min(
                img_display_w / iw,
                max_img_h / ih,
                1.0,
            )
            display_w = min(iw * ratio, img_display_w)
            display_h = min(ih * ratio, max_img_h)

            # Absolute safety cap (should never trigger)
            if display_w <= 0 or display_h <= 0:
                continue
            if display_h > 300:
                display_h = 300
                display_w = display_h * (iw / ih)

            img = RLImage(
                io.BytesIO(compressed),
                width=display_w,
                height=display_h,
            )

            caption = photo.get("caption", "")
            if caption:
                # Use a plain list — KeepTogether inside Table
                # cells causes ReportLab to report infinite height
                cell = [
                    img,
                    Spacer(1, 2),
                    Paragraph(caption, caption_style),
                ]
            else:
                cell = img

            row_cells.append(cell)

            if len(row_cells) == cols:
                rows.append(row_cells)
                row_cells = []
        except Exception:
            continue

    if row_cells:
        while len(row_cells) < cols:
            row_cells.append("")
        rows.append(row_cells)

    if rows:
        photo_table = Table(
            rows,
            colWidths=[cell_width] * cols,
            hAlign="LEFT",
        )
        photo_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(photo_table)
